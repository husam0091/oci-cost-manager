"""Admin routes: login, settings, run scan, runs list."""
from fastapi import APIRouter, Depends, HTTPException, Response, Cookie, Request
from fastapi.responses import FileResponse
from pydantic import BaseModel, Field
from typing import Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import func
from pathlib import Path
from datetime import UTC, datetime, timedelta
import subprocess
import threading
import json
import re
import csv
import uuid
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from core.database import ensure_settings_schema, get_db, SessionLocal
from core.errors import raise_production_block
from core.models import ActionEvent, ActionRequest, AllocationRule, BudgetAlertEvent, Compartment, Setting, ScanRun, Resource, CostSnapshot, TrendPoint, UserAccount
from core.auth import verify_password, hash_password, create_access_token, decode_token
from core.rbac import feature_flags, resolve_principal
from core.scheduler import schedule_scan
from core.config import get_settings as get_app_settings
from services.scanner import run_full_scan
from services.oci_client import reset_oci_client, test_oci_connection
from services import get_cost_calculator
from services.allocation import evaluate_allocation, load_enabled_rules
from services.budget_engine import evaluate_budget_statuses
from services.recommendations import generate_recommendations

router = APIRouter()

_PROD_BLOCKED_FIELDS = {"oci_key_content", "oci_key_file", "oci_pass_phrase", "oci_config_file"}


class LoginRequest(BaseModel):
    username: str
    password: str


class SettingsUpdate(BaseModel):
    username: Optional[str] = None
    password: Optional[str] = None
    scan_interval_hours: Optional[int] = None
    oci_auth_mode: Optional[str] = None
    oci_config_profile: Optional[str] = None
    oci_config_file: Optional[str] = None
    oci_user: Optional[str] = None
    oci_fingerprint: Optional[str] = None
    oci_tenancy: Optional[str] = None
    oci_region: Optional[str] = None
    oci_key_file: Optional[str] = None
    oci_key_content: Optional[str] = None
    oci_pass_phrase: Optional[str] = None
    important_compartment_ids: Optional[list[str]] = None
    important_include_children: Optional[bool] = None
    notifications_email_enabled: Optional[bool] = None
    notifications_smtp_host: Optional[str] = None
    notifications_smtp_port: Optional[int] = None
    notifications_smtp_username: Optional[str] = None
    notifications_smtp_password: Optional[str] = None
    notifications_email_from: Optional[str] = None
    notifications_email_to: Optional[list[str]] = None
    notifications_webhook_enabled: Optional[bool] = None
    notifications_webhook_url: Optional[str] = None
    notifications_webhook_dry_run: Optional[bool] = None
    user_role: Optional[str] = None
    allowed_teams: Optional[list[str]] = None
    allowed_apps: Optional[list[str]] = None
    allowed_envs: Optional[list[str]] = None
    allowed_compartment_ids: Optional[list[str]] = None
    enable_oci_executors: Optional[bool] = None
    enable_destructive_actions: Optional[bool] = None
    enable_budget_auto_eval: Optional[bool] = None
    enable_demo_mode: Optional[bool] = None


class OciConnectionTestRequest(BaseModel):
    oci_auth_mode: Optional[str] = None
    oci_config_profile: Optional[str] = None
    oci_config_file: Optional[str] = None
    oci_user: Optional[str] = None
    oci_fingerprint: Optional[str] = None
    oci_tenancy: Optional[str] = None
    oci_region: Optional[str] = None
    oci_key_file: Optional[str] = None
    oci_key_content: Optional[str] = None
    oci_pass_phrase: Optional[str] = None


class ExportSnapshotRequest(BaseModel):
    name: Optional[str] = None
    report_type: str = "snapshot"  # snapshot|cyber|inventory|cost_validation
    export_format: str = "json"  # json|csv|xlsx
    include_resource_samples: int = 200
    include_scan_runs: bool = True
    include_trend_points: bool = True
    include_cost_snapshot: bool = True


class ImportantCompartmentsUpdate(BaseModel):
    important_compartments: list[str] = Field(default_factory=list)
    include_children: bool = True


class ExportGenerateRequest(BaseModel):
    report_type: str
    start_date: str
    end_date: str
    options: dict[str, Any] = Field(default_factory=dict)


class UserCreateRequest(BaseModel):
    username: str
    password: str
    role: str = "viewer"
    allowed_teams: list[str] = Field(default_factory=list)
    allowed_apps: list[str] = Field(default_factory=list)
    allowed_envs: list[str] = Field(default_factory=list)
    allowed_compartment_ids: list[str] = Field(default_factory=list)
    is_active: bool = True


class UserUpdateRequest(BaseModel):
    password: Optional[str] = None
    role: Optional[str] = None
    allowed_teams: Optional[list[str]] = None
    allowed_apps: Optional[list[str]] = None
    allowed_envs: Optional[list[str]] = None
    allowed_compartment_ids: Optional[list[str]] = None
    is_active: Optional[bool] = None


def _require_admin(token: Optional[str] = Cookie(default=None, alias="access_token"), db: Session = Depends(get_db)):
    ensure_settings_schema()
    try:
        principal = resolve_principal(db, token, strict=True)
    except PermissionError as exc:
        raise HTTPException(status_code=401, detail=str(exc))
    if principal.role != "admin":
        raise HTTPException(status_code=403, detail="Admin role required")
    return {"sub": principal.username}


def _production_blocked_field(payload: BaseModel) -> Optional[str]:
    cfg = get_app_settings()
    if cfg.app_env.lower() != "production":
        return None
    if cfg.allow_oci_file_path_mode:
        return None
    data = payload.model_dump(exclude_none=True)
    for field in _PROD_BLOCKED_FIELDS:
        value = data.get(field)
        if value not in (None, "", []):
            return field
    return None


@router.post("/login")
async def login(req: LoginRequest, resp: Response, db: Session = Depends(get_db)):
    username = (req.username or "").strip()
    user = db.query(UserAccount).filter(UserAccount.username == username, UserAccount.is_active == True).one_or_none()
    if user and verify_password(req.password, user.password_hash):
        token = create_access_token(subject=user.username, expires_minutes=60)
        resp.set_cookie(
            "access_token",
            token,
            httponly=True,
            samesite="lax",
            max_age=3600,
        )
        return {"success": True}

    # Backward compatibility with legacy settings admin login
    s = db.query(Setting).filter(Setting.id == 1).one_or_none()
    if s and username == s.username and verify_password(req.password, s.password_hash):
        if not user:
            db.add(UserAccount(
                username=s.username,
                password_hash=s.password_hash,
                role=(s.user_role or "admin"),
                allowed_teams=list(s.allowed_teams or []),
                allowed_apps=list(s.allowed_apps or []),
                allowed_envs=list(s.allowed_envs or []),
                allowed_compartment_ids=list(s.allowed_compartment_ids or []),
                is_active=True,
            ))
            db.commit()
        token = create_access_token(subject=s.username, expires_minutes=60)
        resp.set_cookie(
            "access_token",
            token,
            httponly=True,
            samesite="lax",
            max_age=3600,
        )
        return {"success": True}

    raise HTTPException(status_code=401, detail="Invalid credentials")


@router.post("/logout")
async def logout(resp: Response):
    resp.delete_cookie("access_token")
    return {"success": True}


@router.get("/settings")
async def get_settings(db: Session = Depends(get_db), user=Depends(_require_admin)):
    ensure_settings_schema()
    s = db.query(Setting).filter(Setting.id == 1).one()
    return {
        "success": True,
        "data": {
            "username": s.username,
            "scan_interval_hours": s.scan_interval_hours,
            "oci_auth_mode": getattr(s, "oci_auth_mode", "profile"),
            "oci_config_profile": getattr(s, "oci_config_profile", "DEFAULT"),
            "oci_user": getattr(s, "oci_user", None),
            "oci_fingerprint": getattr(s, "oci_fingerprint", None),
            "oci_tenancy": getattr(s, "oci_tenancy", None),
            "oci_region": getattr(s, "oci_region", None),
            "oci_last_test_status": getattr(s, "oci_last_test_status", None),
            "oci_last_tested_at": s.oci_last_tested_at.isoformat() if getattr(s, "oci_last_tested_at", None) else None,
            "important_compartment_ids": getattr(s, "important_compartment_ids", None) or [],
            "important_include_children": bool(getattr(s, "important_include_children", True)),
            "notifications_email_enabled": bool(getattr(s, "notifications_email_enabled", False)),
            "notifications_smtp_host": getattr(s, "notifications_smtp_host", None),
            "notifications_smtp_port": getattr(s, "notifications_smtp_port", 587),
            "notifications_smtp_username": getattr(s, "notifications_smtp_username", None),
            "notifications_smtp_password": getattr(s, "notifications_smtp_password", None),
            "notifications_email_from": getattr(s, "notifications_email_from", None),
            "notifications_email_to": list(getattr(s, "notifications_email_to", None) or []),
            "notifications_webhook_enabled": bool(getattr(s, "notifications_webhook_enabled", False)),
            "notifications_webhook_url": getattr(s, "notifications_webhook_url", None),
            "notifications_webhook_dry_run": bool(getattr(s, "notifications_webhook_dry_run", True)),
            "user_role": getattr(s, "user_role", "admin"),
            "allowed_teams": list(getattr(s, "allowed_teams", None) or []),
            "allowed_apps": list(getattr(s, "allowed_apps", None) or []),
            "allowed_envs": list(getattr(s, "allowed_envs", None) or []),
            "allowed_compartment_ids": list(getattr(s, "allowed_compartment_ids", None) or []),
            "feature_flags": feature_flags(s),
            "secrets_note": "Secrets can come from direct value, env var overrides, or vault:// references.",
        },
    }


@router.put("/settings")
async def update_settings(req: SettingsUpdate, request: Request, db: Session = Depends(get_db), user=Depends(_require_admin)):
    ensure_settings_schema()
    blocked = _production_blocked_field(req)
    if blocked:
        raise_production_block(blocked, getattr(request.state, "correlation_id", None))
    s = db.query(Setting).filter(Setting.id == 1).one()
    changed = False
    if req.username:
        s.username = req.username; changed = True
    if req.password:
        s.password_hash = hash_password(req.password); changed = True
    if req.scan_interval_hours and req.scan_interval_hours > 0:
        s.scan_interval_hours = req.scan_interval_hours; changed = True
    if req.oci_auth_mode is not None:
        mode = (req.oci_auth_mode or "").strip().lower()
        if mode in ("profile", "direct"):
            s.oci_auth_mode = mode
            changed = True
    if req.oci_config_profile is not None:
        profile = req.oci_config_profile.strip()
        if profile:
            s.oci_config_profile = profile
            changed = True
    if req.oci_config_file is not None:
        cfg_file = req.oci_config_file.strip()
        s.oci_config_file = cfg_file or None
        changed = True
    if req.oci_user is not None:
        s.oci_user = (req.oci_user or "").strip() or None
        changed = True
    if req.oci_fingerprint is not None:
        s.oci_fingerprint = (req.oci_fingerprint or "").strip() or None
        changed = True
    if req.oci_tenancy is not None:
        s.oci_tenancy = (req.oci_tenancy or "").strip() or None
        changed = True
    if req.oci_region is not None:
        s.oci_region = (req.oci_region or "").strip() or None
        changed = True
    if req.oci_key_file is not None:
        s.oci_key_file = (req.oci_key_file or "").strip() or None
        changed = True
    if req.oci_key_content is not None:
        key_content = (req.oci_key_content or "").strip()
        s.oci_key_content = key_content or None
        changed = True
    if req.oci_pass_phrase is not None:
        s.oci_pass_phrase = (req.oci_pass_phrase or "").strip() or None
        changed = True
    if req.important_compartment_ids is not None:
        s.important_compartment_ids = req.important_compartment_ids
        changed = True
    if req.important_include_children is not None:
        s.important_include_children = bool(req.important_include_children)
        changed = True
    if req.notifications_email_enabled is not None:
        s.notifications_email_enabled = bool(req.notifications_email_enabled)
        changed = True
    if req.notifications_smtp_host is not None:
        s.notifications_smtp_host = (req.notifications_smtp_host or "").strip() or None
        changed = True
    if req.notifications_smtp_port is not None:
        s.notifications_smtp_port = int(req.notifications_smtp_port)
        changed = True
    if req.notifications_smtp_username is not None:
        s.notifications_smtp_username = (req.notifications_smtp_username or "").strip() or None
        changed = True
    if req.notifications_smtp_password is not None:
        s.notifications_smtp_password = req.notifications_smtp_password or None
        changed = True
    if req.notifications_email_from is not None:
        s.notifications_email_from = (req.notifications_email_from or "").strip() or None
        changed = True
    if req.notifications_email_to is not None:
        s.notifications_email_to = [x.strip() for x in (req.notifications_email_to or []) if x and x.strip()]
        changed = True
    if req.notifications_webhook_enabled is not None:
        s.notifications_webhook_enabled = bool(req.notifications_webhook_enabled)
        changed = True
    if req.notifications_webhook_url is not None:
        s.notifications_webhook_url = (req.notifications_webhook_url or "").strip() or None
        changed = True
    if req.notifications_webhook_dry_run is not None:
        s.notifications_webhook_dry_run = bool(req.notifications_webhook_dry_run)
        changed = True
    if req.user_role is not None:
        role = (req.user_role or "").strip().lower()
        if role in {"admin", "finops", "engineer", "viewer"}:
            s.user_role = role
            changed = True
    if req.allowed_teams is not None:
        s.allowed_teams = [x.strip() for x in (req.allowed_teams or []) if x and x.strip()]
        changed = True
    if req.allowed_apps is not None:
        s.allowed_apps = [x.strip() for x in (req.allowed_apps or []) if x and x.strip()]
        changed = True
    if req.allowed_envs is not None:
        s.allowed_envs = [x.strip() for x in (req.allowed_envs or []) if x and x.strip()]
        changed = True
    if req.allowed_compartment_ids is not None:
        s.allowed_compartment_ids = [x.strip() for x in (req.allowed_compartment_ids or []) if x and x.strip()]
        changed = True
    if req.enable_oci_executors is not None:
        s.enable_oci_executors = bool(req.enable_oci_executors)
        changed = True
    if req.enable_destructive_actions is not None:
        s.enable_destructive_actions = bool(req.enable_destructive_actions)
        changed = True
    if req.enable_budget_auto_eval is not None:
        s.enable_budget_auto_eval = bool(req.enable_budget_auto_eval)
        changed = True
    if req.enable_demo_mode is not None:
        s.enable_demo_mode = bool(req.enable_demo_mode)
        changed = True
    if changed:
        db.commit()
        reset_oci_client()
        schedule_scan()
    return {"success": True}


@router.get("/settings/feature-flags")
async def get_feature_flags(db: Session = Depends(get_db), user=Depends(_require_admin)):
    s = db.query(Setting).filter(Setting.id == 1).one_or_none()
    return {"success": True, "data": feature_flags(s)}


@router.post("/settings/feature-flags")
async def update_feature_flags(payload: dict[str, Any], db: Session = Depends(get_db), user=Depends(_require_admin)):
    s = db.query(Setting).filter(Setting.id == 1).one()
    if "enable_oci_executors" in payload:
        s.enable_oci_executors = bool(payload.get("enable_oci_executors"))
    if "enable_destructive_actions" in payload:
        s.enable_destructive_actions = bool(payload.get("enable_destructive_actions"))
    if "enable_budget_auto_eval" in payload:
        s.enable_budget_auto_eval = bool(payload.get("enable_budget_auto_eval"))
    if "enable_demo_mode" in payload:
        s.enable_demo_mode = bool(payload.get("enable_demo_mode"))
    db.commit()
    return {"success": True, "data": feature_flags(s)}


@router.get("/settings/important-compartments")
async def get_important_compartments(db: Session = Depends(get_db), user=Depends(_require_admin)):
    s = db.query(Setting).filter(Setting.id == 1).one()
    ids = list(getattr(s, "important_compartments", None) or getattr(s, "important_compartment_ids", None) or [])
    include_children = bool(getattr(s, "important_include_children", True))
    all_comps = db.query(Compartment).all()
    if not ids:
        ids = [c.id for c in all_comps if c.name and c.name.lower() in {"foo", "ad1"}]
    return {
        "success": True,
        "data": {
            "important_compartments": ids,
            "include_children": include_children,
        },
    }


@router.post("/settings/important-compartments")
async def set_important_compartments(
    req: ImportantCompartmentsUpdate,
    db: Session = Depends(get_db),
    user=Depends(_require_admin),
):
    s = db.query(Setting).filter(Setting.id == 1).one()
    s.important_compartments = req.important_compartments or []
    s.important_compartment_ids = req.important_compartments or []
    s.important_include_children = bool(req.include_children)
    db.commit()
    return {
        "success": True,
        "data": {
            "important_compartments": list(s.important_compartments or s.important_compartment_ids or []),
            "include_children": s.important_include_children,
        },
    }


@router.post("/settings/test-oci")
async def test_settings_oci(req: OciConnectionTestRequest, request: Request, db: Session = Depends(get_db), user=Depends(_require_admin)):
    blocked = _production_blocked_field(req)
    if blocked:
        raise_production_block(blocked, getattr(request.state, "correlation_id", None))
    s = db.query(Setting).filter(Setting.id == 1).one()
    runtime = {
        "auth_mode": req.oci_auth_mode or getattr(s, "oci_auth_mode", "profile"),
        "config_profile": req.oci_config_profile if req.oci_config_profile is not None else getattr(s, "oci_config_profile", "DEFAULT"),
        "config_file": req.oci_config_file if req.oci_config_file is not None else getattr(s, "oci_config_file", None),
        "user": req.oci_user if req.oci_user is not None else getattr(s, "oci_user", None),
        "fingerprint": req.oci_fingerprint if req.oci_fingerprint is not None else getattr(s, "oci_fingerprint", None),
        "tenancy": req.oci_tenancy if req.oci_tenancy is not None else getattr(s, "oci_tenancy", None),
        "region": req.oci_region if req.oci_region is not None else getattr(s, "oci_region", None),
        "key_file": req.oci_key_file if req.oci_key_file is not None else getattr(s, "oci_key_file", None),
        "key_content": req.oci_key_content if req.oci_key_content is not None else getattr(s, "oci_key_content", None),
        "pass_phrase": req.oci_pass_phrase if req.oci_pass_phrase is not None else getattr(s, "oci_pass_phrase", None),
    }
    try:
        data = test_oci_connection(runtime)
        s.oci_last_test_status = "healthy"
        s.oci_last_tested_at = datetime.now(UTC)
        s.oci_last_test_error = None
        db.commit()
        return {"success": True, "data": data}
    except Exception as exc:
        correlation_id = getattr(request.state, "correlation_id", None) or str(uuid.uuid4())
        s.oci_last_test_status = "failed"
        s.oci_last_tested_at = datetime.now(UTC)
        s.oci_last_test_error = "redacted"
        db.commit()
        raise HTTPException(
            status_code=400,
            detail={
                "success": False,
                "error": {
                    "code": "OCI_TEST_FAILED",
                    "reason": "redacted",
                    "correlation_id": correlation_id,
                },
            },
        )


@router.post("/scan/run")
async def scan_now(db: Session = Depends(get_db), user=Depends(_require_admin)):
    _mark_stale_scan_runs(db)

    running = db.query(ScanRun).filter(ScanRun.status == "running").order_by(ScanRun.id.desc()).first()
    if running:
        return {"success": True, "data": {"status": "already_running", "run_id": running.id}}

    def _run_scan_in_background():
        local_db = SessionLocal()
        try:
            run_full_scan(local_db)
        finally:
            local_db.close()

    t = threading.Thread(target=_run_scan_in_background, daemon=True)
    t.start()
    return {"success": True, "data": {"status": "started"}}


@router.get("/scan/runs")
async def list_runs(db: Session = Depends(get_db), user=Depends(_require_admin)):
    _mark_stale_scan_runs(db)
    runs = db.query(ScanRun).order_by(ScanRun.id.desc()).limit(20).all()
    out = [{
        "id": r.id, "status": r.status, "started_at": r.started_at.isoformat() if r.started_at else None,
        "finished_at": r.finished_at.isoformat() if r.finished_at else None, "error_message": r.error_message
    } for r in runs]
    return {"success": True, "data": out}


def _mark_stale_scan_runs(db: Session) -> None:
    """Mark orphaned long-running scans as failed to avoid indefinite UI scanning."""
    stale_cutoff = datetime.now(UTC) - timedelta(minutes=30)
    stale_runs = db.query(ScanRun).filter(ScanRun.status == "running", ScanRun.started_at < stale_cutoff).all()
    for run in stale_runs:
        run.status = "failed"
        run.finished_at = datetime.now(UTC)
        run.error_message = "Marked stale after 30m timeout"
    if stale_runs:
        db.commit()


@router.post("/exports/snapshot")
async def export_snapshot(req: ExportSnapshotRequest, db: Session = Depends(get_db), user=Depends(_require_admin)):
    s = db.query(Setting).filter(Setting.id == 1).one()
    now = datetime.now(UTC)
    counts = db.query(Resource.type, func.count(Resource.id)).group_by(Resource.type).all()
    latest_monthly = db.query(CostSnapshot).filter(CostSnapshot.period == "monthly").order_by(CostSnapshot.start_date.desc()).first() if req.include_cost_snapshot else None
    trends = db.query(TrendPoint).order_by(TrendPoint.month_start.desc()).limit(6).all() if req.include_trend_points else []
    recent_runs = db.query(ScanRun).order_by(ScanRun.id.desc()).limit(20).all() if req.include_scan_runs else []
    latest_scan = db.query(ScanRun).order_by(ScanRun.id.desc()).first()
    sample_limit = max(0, min(req.include_resource_samples, 2000))
    sample_resources = db.query(Resource).order_by(Resource.updated_at.desc()).limit(sample_limit).all()

    safe_name = re.sub(r"[^a-zA-Z0-9_-]", "-", (req.name or "").strip()).strip("-")
    stamp = now.strftime("%Y%m%d-%H%M%S")
    fmt = (req.export_format or "json").lower()
    if fmt not in {"json", "csv", "xlsx"}:
        raise HTTPException(status_code=400, detail="export_format must be one of: json, csv, xlsx")
    report_type = (req.report_type or "snapshot").lower()
    if report_type not in {"snapshot", "cyber", "inventory", "cost_validation"}:
        raise HTTPException(status_code=400, detail="report_type must be one of: snapshot, cyber, inventory, cost_validation")
    filename = f"export-report-{safe_name + '-' if safe_name else ''}{stamp}.{fmt}"

    cfg = get_app_settings()
    export_dir = Path(cfg.export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    out_path = export_dir / filename

    payload = {
        "meta": {
            "generated_at": now.isoformat(),
            "generated_by": user.get("sub"),
            "actor": user.get("sub"),
            "export_name": req.name,
            "source": "oci-cost-manager",
            "version": "1.0.0",
            "format": fmt,
            "report_type": report_type,
            "oci_auth_mode": getattr(s, "oci_auth_mode", "profile"),
            "oci_config_profile": getattr(s, "oci_config_profile", None),
            "scan_run_id": latest_scan.id if latest_scan else None,
        },
        "integration": {
            "auth_mode": getattr(s, "oci_auth_mode", "profile"),
            "profile": getattr(s, "oci_config_profile", None),
            "region": getattr(s, "oci_region", None),
        },
        "resource_counts": {t: c for t, c in counts},
        "validation": {
            "has_cost_snapshot": latest_monthly is not None,
            "has_trends": len(trends) > 0,
            "has_scan_runs": len(recent_runs) > 0,
            "resource_sample_count": len(sample_resources),
            "resource_types_count": len(counts),
        },
        "latest_monthly_cost": {
            "period": latest_monthly.period if latest_monthly else None,
            "start_date": latest_monthly.start_date.isoformat() if latest_monthly else None,
            "end_date": latest_monthly.end_date.isoformat() if latest_monthly and latest_monthly.end_date else None,
            "total": latest_monthly.total if latest_monthly else 0,
            "by_service": latest_monthly.by_service if latest_monthly else {},
        },
        "trend_points": [
            {"month": t.month, "total_cost": t.total_cost, "by_service": t.by_service or {}}
            for t in reversed(trends)
        ],
        "scan_runs": [
            {
                "id": r.id,
                "status": r.status,
                "started_at": r.started_at.isoformat() if r.started_at else None,
                "finished_at": r.finished_at.isoformat() if r.finished_at else None,
                "error_message": r.error_message,
            }
            for r in recent_runs
        ],
        "resource_samples": [
            {
                "ocid": r.ocid,
                "name": r.name,
                "type": r.type,
                "compartment_id": r.compartment_id,
                "status": r.status,
                "shape": r.shape,
                "details": r.details or {},
            }
            for r in sample_resources
        ],
    }

    if fmt == "json":
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2)
    elif fmt == "csv":
        _write_export_csv(out_path, payload, report_type)
    else:
        if report_type == "cyber":
            _write_export_xlsx_cyber(out_path, payload)
        elif report_type == "inventory":
            _write_export_xlsx_inventory(out_path, payload)
        elif report_type == "cost_validation":
            _write_export_xlsx_cost_validation(out_path, payload)
        else:
            _write_export_xlsx(out_path, payload)

    return {
        "success": True,
        "data": {
            "file_name": filename,
            "file_path": str(out_path),
            "format": fmt,
            "report_type": report_type,
            "resource_sample_count": len(payload["resource_samples"]),
        },
    }


@router.get("/exports/list")
async def list_exports(user=Depends(_require_admin)):
    cfg = get_app_settings()
    export_dir = Path(cfg.export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    files = sorted(
        [p for p in export_dir.iterdir() if p.is_file()],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )[:100]
    out = []
    for p in files:
        entry = {
            "name": p.name,
            "path": str(p),
            "download_url": f"/api/v1/admin/exports/download/{p.name}",
            "size_bytes": p.stat().st_size,
            "updated_at": datetime.fromtimestamp(p.stat().st_mtime, UTC).isoformat(),
            "report_type": None,
            "range": None,
        }
        if p.name.endswith(".xlsx"):
            manifest = p.with_name(p.name.replace(".xlsx", ".manifest.json"))
            if manifest.exists():
                try:
                    meta = json.loads(manifest.read_text(encoding="utf-8"))
                    entry["report_type"] = meta.get("report_type")
                    entry["range"] = {"start_date": meta.get("start_date"), "end_date": meta.get("end_date")}
                    entry["manifest_url"] = f"/api/v1/admin/exports/download/{manifest.name}"
                    validation = p.with_name(p.name.replace(".xlsx", ".validation.json"))
                    if validation.exists():
                        entry["validation_url"] = f"/api/v1/admin/exports/download/{validation.name}"
                except Exception:
                    pass
        out.append(entry)
    return {"success": True, "data": out}


@router.get("/exports/download/{name}")
async def download_export(name: str, user=Depends(_require_admin)):
    if "/" in name or "\\" in name or ".." in name:
        raise HTTPException(status_code=400, detail="Invalid file name")
    cfg = get_app_settings()
    path = Path(cfg.export_dir) / name
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="Export not found")
    suffix = path.suffix.lower()
    media_map = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".csv": "text/csv",
        ".json": "application/json",
    }
    media_type = media_map.get(suffix, "application/octet-stream")
    return FileResponse(path=path, filename=name, media_type=media_type)


def _write_export_csv(path: Path, payload: dict, report_type: str) -> None:
    fields = [
        "ocid", "name", "type", "compartment_id", "status", "shape",
        "image_name", "image_family", "image_vendor", "attachment_state",
        "size_display", "protocol",
    ]
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        for r in payload.get("resource_samples", []):
            d = r.get("details") or {}
            writer.writerow({
                "ocid": r.get("ocid"),
                "name": r.get("name"),
                "type": r.get("type"),
                "compartment_id": r.get("compartment_id"),
                "status": r.get("status"),
                "shape": r.get("shape"),
                "image_name": d.get("image_name"),
                "image_family": d.get("image_family"),
                "image_vendor": d.get("image_vendor"),
                "attachment_state": d.get("attachment_state"),
                "size_display": d.get("size_display"),
                "protocol": d.get("protocol"),
            })
        if report_type == "cost_validation":
            writer.writerow({k: "" for k in fields})
            writer.writerow({"ocid": "VALIDATION", "name": "has_cost_snapshot", "type": str(payload.get("validation", {}).get("has_cost_snapshot"))})
            writer.writerow({"ocid": "VALIDATION", "name": "has_trends", "type": str(payload.get("validation", {}).get("has_trends"))})
            writer.writerow({"ocid": "VALIDATION", "name": "has_scan_runs", "type": str(payload.get("validation", {}).get("has_scan_runs"))})


def _write_export_xlsx(path: Path, payload: dict) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Generated At", payload["meta"].get("generated_at")])
    ws_summary.append(["Generated By", payload["meta"].get("generated_by")])
    ws_summary.append(["Auth Mode", payload.get("integration", {}).get("auth_mode")])
    ws_summary.append(["Region", payload.get("integration", {}).get("region")])
    ws_summary.append(["Monthly Total Cost", payload.get("latest_monthly_cost", {}).get("total")])

    ws_counts = wb.create_sheet("Resource Counts")
    ws_counts.append(["Type", "Count"])
    for t, c in (payload.get("resource_counts") or {}).items():
        ws_counts.append([t, c])

    ws_trends = wb.create_sheet("Cost Trends")
    ws_trends.append(["Month", "Total Cost"])
    for t in payload.get("trend_points", []):
        ws_trends.append([t.get("month"), t.get("total_cost")])

    ws_runs = wb.create_sheet("Scan Runs")
    ws_runs.append(["Run ID", "Status", "Started", "Finished", "Error"])
    for r in payload.get("scan_runs", []):
        ws_runs.append([r.get("id"), r.get("status"), r.get("started_at"), r.get("finished_at"), r.get("error_message")])

    ws_resources = wb.create_sheet("Resource Samples")
    headers = [
        "OCID", "Name", "Type", "Compartment", "Status", "Shape",
        "Image Name", "Image Family", "Image Vendor", "Attachment State", "Size", "Protocol",
    ]
    ws_resources.append(headers)
    for r in payload.get("resource_samples", []):
        d = r.get("details") or {}
        ws_resources.append([
            r.get("ocid"),
            r.get("name"),
            r.get("type"),
            r.get("compartment_id"),
            r.get("status"),
            r.get("shape"),
            d.get("image_name"),
            d.get("image_family"),
            d.get("image_vendor"),
            d.get("attachment_state"),
            d.get("size_display"),
            d.get("protocol"),
        ])
    wb.save(path)


def _write_export_xlsx_cyber(path: Path, payload: dict) -> None:
    """Write workbook with sheet layout similar to example/oci_report_cyber.py."""
    wb = Workbook()
    ws_inventory = wb.active
    ws_inventory.title = "OCI_Instance_Inventory"
    ws_inventory.append([
        "Instance Name", "Instance OCID", "Type", "State", "Shape",
        "Operating System", "License Model", "Image Vendor", "Attachment", "Storage Size",
    ])
    resources = payload.get("resource_samples", [])
    for r in resources:
        d = r.get("details") or {}
        ws_inventory.append([
            r.get("name"),
            r.get("ocid"),
            r.get("type"),
            r.get("status"),
            r.get("shape"),
            d.get("image_name"),
            d.get("license_model"),
            d.get("image_vendor"),
            d.get("attachment_state"),
            d.get("size_display"),
        ])

    ws_windows = wb.create_sheet("Windows_License_Calculation")
    ws_windows.append(["Instance", "OS", "Estimated Cost Component (USD)"])
    monthly = payload.get("latest_monthly_cost", {}).get("by_service", {}) or {}
    windows_estimate = 0.0
    for k, v in monthly.items():
        if "windows" in str(k).lower() or "microsoft" in str(k).lower():
            windows_estimate += float(v or 0)
    ws_windows.append(["Windows/Microsoft services", "Windows", round(windows_estimate, 4)])

    ws_summary = wb.create_sheet("Cost_Summary_(USD_SAR)")
    ws_summary.append(["Metric", "USD", "SAR (3.75x)"])
    total = float(payload.get("latest_monthly_cost", {}).get("total") or 0)
    ws_summary.append(["Latest Monthly Total", round(total, 4), round(total * 3.75, 4)])
    ws_summary.append(["Windows-related estimate", round(windows_estimate, 4), round(windows_estimate * 3.75, 4)])
    ws_summary.append(["Final COST", round(total + windows_estimate, 4), round((total + windows_estimate) * 3.75, 4)])

    ws_full_year = wb.create_sheet("Full-Year Estimated Costs (USD)")
    ws_full_year.append(["Metric", "USD"])
    ws_full_year.append(["Estimated Full Year", round(total * 12, 4)])
    ws_full_year.append(["Estimated Full Year + Windows", round((total + windows_estimate) * 12, 4)])

    wb.save(path)


def _write_export_xlsx_inventory(path: Path, payload: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["Name", "OCID", "Type", "Status", "Compartment", "Image", "Vendor", "Attachment", "Size"])
    for r in payload.get("resource_samples", []):
        d = r.get("details") or {}
        ws.append([
            r.get("name"),
            r.get("ocid"),
            r.get("type"),
            r.get("status"),
            r.get("compartment_id"),
            d.get("image_name"),
            d.get("image_vendor"),
            d.get("attachment_state"),
            d.get("size_display"),
        ])
    wb.save(path)


def _write_export_xlsx_cost_validation(path: Path, payload: dict) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation"
    ws.append(["Check", "Value"])
    for k, v in (payload.get("validation") or {}).items():
        ws.append([k, str(v)])

    ws_cost = wb.create_sheet("Monthly Cost")
    ws_cost.append(["Service", "Cost"])
    by_service = payload.get("latest_monthly_cost", {}).get("by_service") or {}
    for s, c in by_service.items():
        ws_cost.append([s, c])
    ws_cost.append(["TOTAL", payload.get("latest_monthly_cost", {}).get("total") or 0])
    wb.save(path)


def _parse_iso_date(raw: str, *, is_end: bool) -> datetime:
    try:
        if "T" in raw:
            dt = datetime.fromisoformat(raw.replace("Z", "+00:00"))
        else:
            dt = datetime.fromisoformat(raw).replace(hour=0, minute=0, second=0, microsecond=0)
            if is_end:
                dt = dt + timedelta(days=1)
    except ValueError:
        raise HTTPException(status_code=422, detail=f"Invalid date: {raw}")
    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC)
    return dt.astimezone(UTC)


def _safe_pct(delta: float, base: float) -> float:
    return (delta / base * 100.0) if base else 0.0


def _resource_tag(details: dict, *keys: str) -> str:
    tags = (details or {}).get("defined_tags") or (details or {}).get("freeform_tags") or {}
    for key in keys:
        if tags.get(key):
            return str(tags.get(key))
    return "Unallocated"


def _report_catalog() -> dict[str, dict]:
    return {
        "executive_summary_monthly": {"title": "Executive Summary", "audience": "Exec/Finance", "tier": 1},
        "cost_by_service": {"title": "Cost by Service", "audience": "Finance/Ops", "tier": 1},
        "cost_by_compartment": {"title": "Cost by Compartment", "audience": "Finance/Ops", "tier": 1},
        "top_resources_by_cost": {"title": "Top Resources by Cost", "audience": "FinOps/Ops", "tier": 1},
        "mapping_health_unallocated": {"title": "Unallocated / Low-confidence Mapping", "audience": "Finance/FinOps", "tier": 1},
        "showback_team_app_env": {"title": "Showback by Team/App/Env", "audience": "Finance/FinOps", "tier": 1},
        "inventory_summary_by_compartment": {"title": "Inventory Summary by Compartment", "audience": "Ops", "tier": 2},
        "storage_backup_governance": {"title": "Storage & Backup Governance", "audience": "Ops/Governance", "tier": 2},
        "license_spend": {"title": "License Spend", "audience": "Finance/Governance", "tier": 2},
        "movers_and_anomalies": {"title": "Movers & Anomalies", "audience": "Exec/Ops", "tier": 2},
        "optimization_recommendations": {"title": "Optimization Recommendations", "audience": "Exec/FinOps/Ops", "tier": 2},
        "budget_health": {"title": "Budget Health", "audience": "Exec/Finance/Engineering", "tier": 2},
        "actions_audit": {"title": "Actions Audit", "audience": "Exec/FinOps/Ops/Sec", "tier": 2},
        "ops_audit": {"title": "Ops Audit", "audience": "Platform/Ops", "tier": 2},
        "vm_windows_inventory": {"title": "VM & Windows Server Inventory", "audience": "Ops/Security", "tier": 1},
    }


def _legacy_report_map() -> dict[str, str]:
    return {
        "showback_team_app_env": "showback_team_app_env",
        "inventory_summary_compartment": "inventory_summary_by_compartment",
        "anomaly_movers": "movers_and_anomalies",
    }


def _normalize_report_type(report_type: str) -> str:
    normalized = (report_type or "").strip().lower()
    return _legacy_report_map().get(normalized, normalized)


def _short_ocid(value: Optional[str]) -> str:
    if not value:
        return "unknown"
    parts = str(value).split(".")
    return parts[-1][-16:] if parts else str(value)[-16:]


def _daily(cost: float, days: int) -> float:
    return round(cost / max(days, 1), 4)


def _extract_tags(details: dict) -> dict[str, Any]:
    details = details or {}
    tags = {}
    defined = details.get("defined_tags") or {}
    freeform = details.get("freeform_tags") or {}
    if isinstance(defined, dict):
        tags.update(defined)
    if isinstance(freeform, dict):
        tags.update(freeform)
    return tags


def _tag(tags: dict[str, Any], details: dict, *keys: str) -> str:
    for key in keys:
        if key in tags and tags[key]:
            return str(tags[key])
    for key in keys:
        if details.get(key):
            return str(details.get(key))
    return "Unallocated"


def _safe_name(resource: Optional[Resource], fallback: Optional[str]) -> str:
    if resource and resource.name:
        return resource.name
    return _short_ocid(fallback)


def _safe_pct_rounded(delta: float, base: float) -> float:
    return round(_safe_pct(delta, base), 4)


def _license_bucket(sku_name: str, image_name: str, rtype: str) -> str:
    text = (sku_name or "").lower()
    image = (image_name or "").lower()
    rtype = (rtype or "").lower()
    if "sql server" in text or "microsoft sql" in text or "sql_server" == rtype:
        return "sql"
    if "windows" in text or "windows" in image or rtype == "windows_server":
        return "windows"
    if "oracle linux" in text or "oracle os" in text or "oracle linux" in image:
        return "oracle_os"
    return "other"


def _compute_match_confidence(tags: dict[str, Any]) -> tuple[str, str]:
    team = tags.get("owner_team") or tags.get("team")
    app = tags.get("application") or tags.get("app")
    env = tags.get("environment") or tags.get("env")
    present = sum(1 for v in [team, app, env] if v)
    if present >= 3:
        return "high", "team+app+env_tags"
    if present >= 2:
        return "medium", "partial_owner_tags"
    return "low", "missing_owner_tags"


def _get_git_commit() -> Optional[str]:
    try:
        out = subprocess.check_output(["git", "rev-parse", "--short", "HEAD"], text=True, timeout=2)
        return out.strip() or None
    except Exception:
        return None


def _sheet_name(raw: str) -> str:
    safe = re.sub(r"[\[\]\*\?\/\\:]", "_", raw)
    return safe[:31] if len(safe) > 31 else safe


def _build_report_data(
    report_type: str,
    start: datetime,
    end: datetime,
    options: dict,
    db: Session,
    user_name: str,
) -> dict:
    calc = get_cost_calculator()
    prev_end = start
    prev_start = prev_end - (end - start)
    current_services = calc.get_costs_by_service(start, end)
    previous_services = calc.get_costs_by_service(prev_start, prev_end)
    current_rows = calc.get_costs_by_resource(start, end, include_skus=True)
    previous_rows = calc.get_costs_by_resource(prev_start, prev_end, include_skus=True)
    prev_map = {r.get("resource_id"): float(r.get("total_cost") or 0.0) for r in previous_rows}

    resources = {r.ocid: r for r in db.query(Resource).all()}
    compartments = {c.id: c for c in db.query(Compartment).all()}
    latest_scan = db.query(ScanRun).order_by(ScanRun.id.desc()).first()
    app_cfg = get_app_settings()

    summary = {
        "report_type": report_type,
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "generated_at": datetime.now(UTC).isoformat(),
        "generated_by": user_name,
        "scan_id": latest_scan.id if latest_scan else None,
        "total_current": round(sum(current_services.values()), 4),
        "total_previous": round(sum(previous_services.values()), 4),
        "delta_abs": round(sum(current_services.values()) - sum(previous_services.values()), 4),
        "delta_pct": round(_safe_pct(sum(current_services.values()) - sum(previous_services.values()), sum(previous_services.values())), 4),
    }

    data_rows: list[dict] = []
    if report_type == "executive_summary_monthly":
        for svc, cur in sorted(current_services.items(), key=lambda x: x[1], reverse=True)[:5]:
            prev = float(previous_services.get(svc, 0.0))
            data_rows.append({
                "section": "top_driver_service",
                "entity": svc,
                "current_cost": round(cur, 4),
                "previous_cost": round(prev, 4),
                "delta_abs": round(cur - prev, 4),
                "delta_pct": round(_safe_pct(cur - prev, prev), 4),
            })
        movers = sorted(
            [
                {
                    "section": "top_mover_resource",
                    "entity": rid,
                    "current_cost": round(float(row.get("total_cost") or 0.0), 4),
                    "previous_cost": round(float(prev_map.get(rid, 0.0)), 4),
                    "delta_abs": round(float(row.get("total_cost") or 0.0) - float(prev_map.get(rid, 0.0)), 4),
                    "delta_pct": round(_safe_pct(float(row.get("total_cost") or 0.0) - float(prev_map.get(rid, 0.0)), float(prev_map.get(rid, 0.0))), 4),
                }
                for rid, row in [(r.get("resource_id"), r) for r in current_rows if r.get("resource_id")]
            ],
            key=lambda x: abs(x["delta_abs"]),
            reverse=True,
        )[:10]
        data_rows.extend(movers)
    elif report_type == "cost_by_compartment":
        comp_totals: dict[str, float] = {}
        comp_prev: dict[str, float] = {}
        for row in current_rows:
            rid = row.get("resource_id")
            c = resources.get(rid).compartment_id if resources.get(rid) else row.get("compartment_id")
            comp_totals[c] = comp_totals.get(c, 0.0) + float(row.get("total_cost") or 0.0)
            comp_prev[c] = comp_prev.get(c, 0.0) + float(prev_map.get(rid, 0.0))
        for cid, total in sorted(comp_totals.items(), key=lambda x: x[1], reverse=True):
            prev = comp_prev.get(cid, 0.0)
            data_rows.append({
                "compartment_id": cid,
                "compartment_name": compartments.get(cid).name if compartments.get(cid) else cid,
                "current_cost": round(total, 4),
                "previous_cost": round(prev, 4),
                "delta_abs": round(total - prev, 4),
                "delta_pct": round(_safe_pct(total - prev, prev), 4),
                "top_service": max(current_services, key=current_services.get) if current_services else "N/A",
            })
    elif report_type == "showback_team_app_env":
        grouped: dict[tuple[str, str, str], float] = {}
        for row in current_rows:
            rid = row.get("resource_id")
            r = resources.get(rid)
            details = (r.details or {}) if r else {}
            team = _resource_tag(details, "owner_team", "team")
            app = _resource_tag(details, "application", "app")
            env = _resource_tag(details, "environment", "env")
            grouped[(team, app, env)] = grouped.get((team, app, env), 0.0) + float(row.get("total_cost") or 0.0)
        for (team, app, env), total in sorted(grouped.items(), key=lambda x: x[1], reverse=True):
            match_reason = "tag_match" if team != "Unallocated" or app != "Unallocated" or env != "Unallocated" else "fallback"
            data_rows.append({
                "team": team,
                "application": app,
                "environment": env,
                "cost": round(total, 4),
                "match_reason": match_reason,
                "match_confidence": "high" if match_reason == "tag_match" else "low",
            })
    elif report_type == "inventory_summary_compartment":
        category_totals: dict[tuple[str, str], dict[str, float]] = {}
        for row in current_rows:
            rid = row.get("resource_id")
            r = resources.get(rid)
            cid = r.compartment_id if r else (row.get("compartment_id") or "unknown")
            rtype = (r.type if r else "unknown") or "unknown"
            key = (cid, rtype)
            slot = category_totals.setdefault(key, {"count": 0.0, "monthly": 0.0})
            slot["count"] += 1.0
            slot["monthly"] += float(row.get("total_cost") or 0.0)
        for (cid, rtype), vals in sorted(category_totals.items(), key=lambda x: x[1]["monthly"], reverse=True):
            data_rows.append({
                "compartment_id": cid,
                "compartment_name": compartments.get(cid).name if compartments.get(cid) else cid,
                "category": rtype,
                "resource_count": int(vals["count"]),
                "cost_daily": round(vals["monthly"] / max((end - start).days, 1), 4),
                "cost_monthly": round(vals["monthly"], 4),
            })
    elif report_type == "storage_backup_governance":
        storage = {"block": 0.0, "boot": 0.0, "backup": 0.0, "object": 0.0}
        unattached_count = 0
        unattached_cost = 0.0
        for row in current_rows:
            rid = row.get("resource_id")
            r = resources.get(rid)
            total = float(row.get("total_cost") or 0.0)
            if r and (r.details or {}).get("attachment_state") == "UNATTACHED":
                unattached_count += 1
                unattached_cost += total
            for s in row.get("skus") or []:
                sku = (s.get("sku_name") or "").lower()
                cost = float(s.get("cost") or 0.0)
                if "backup" in sku or "snapshot" in sku:
                    storage["backup"] += cost
                elif "boot volume" in sku:
                    storage["boot"] += cost
                elif "block volume" in sku:
                    storage["block"] += cost
                elif "object storage" in sku:
                    storage["object"] += cost
        data_rows = [
            {"metric": "unattached_volume_count", "value": unattached_count},
            {"metric": "unattached_volume_monthly_waste", "value": round(unattached_cost, 4)},
            {"metric": "block_volume_cost", "value": round(storage["block"], 4)},
            {"metric": "boot_volume_cost", "value": round(storage["boot"], 4)},
            {"metric": "backup_snapshot_cost", "value": round(storage["backup"], 4)},
            {"metric": "object_storage_cost", "value": round(storage["object"], 4)},
        ]
    elif report_type == "license_spend":
        for row in current_rows:
            rid = row.get("resource_id")
            r = resources.get(rid)
            for s in row.get("skus") or []:
                sku = (s.get("sku_name") or "").lower()
                if "windows os" in sku or "microsoft sql enterprise" in sku or "sql server" in sku:
                    data_rows.append({
                        "resource_name": r.name if r else rid,
                        "resource_type": r.type if r else "unknown",
                        "sku_name": s.get("sku_name"),
                        "cost": round(float(s.get("cost") or 0.0), 4),
                        "delta_vs_previous": round(float(row.get("total_cost") or 0.0) - float(prev_map.get(rid, 0.0)), 4),
                    })
    elif report_type == "anomaly_movers":
        movers = []
        for row in current_rows:
            rid = row.get("resource_id")
            cur = float(row.get("total_cost") or 0.0)
            prv = float(prev_map.get(rid, 0.0))
            delta = cur - prv
            movers.append({
                "entity": resources.get(rid).name if resources.get(rid) else rid,
                "current": round(cur, 4),
                "previous": round(prv, 4),
                "delta_abs": round(delta, 4),
                "delta_pct": round(_safe_pct(delta, prv), 4),
            })
        data_rows = sorted(movers, key=lambda x: abs(x["delta_abs"]), reverse=True)[:20]
    else:
        raise HTTPException(status_code=400, detail="Unknown report_type")

    missing_mapping = 0
    missing_tags = 0
    for row in current_rows:
        rid = row.get("resource_id")
        r = resources.get(rid)
        if not r:
            missing_mapping += 1
            continue
        tags = (r.details or {}).get("defined_tags") or (r.details or {}).get("freeform_tags") or {}
        if not tags.get("owner_team") or not tags.get("environment"):
            missing_tags += 1

    validation = {
        "row_count": len(data_rows),
        "total_cost_checksum": round(sum(float((r.get("cost") or r.get("current_cost") or r.get("value") or 0.0)) for r in data_rows), 6),
        "missing_mappings_pct": round((missing_mapping / max(len(current_rows), 1)) * 100.0, 4),
        "missing_tags_pct": round((missing_tags / max(len(current_rows), 1)) * 100.0, 4),
    }
    manifest = {
        "report_type": report_type,
        "filters": options or {},
        "start_date": start.isoformat(),
        "end_date": end.isoformat(),
        "scan_id": latest_scan.id if latest_scan else None,
        "generated_at": datetime.now(UTC).isoformat(),
        "app_version": app_cfg.app_version,
        "app_name": app_cfg.app_name,
    }
    return {"summary": summary, "rows": data_rows, "manifest": manifest, "validation": validation}


def _build_report_data_v2(
    report_type: str,
    start: datetime,
    end: datetime,
    start_date_iso: str,
    end_date_iso: str,
    compare: str,
    options: dict,
    db: Session,
    user_name: str,
) -> dict:
    calc = get_cost_calculator()
    prev_end = start
    prev_start = prev_end - (end - start)
    days = max((end - start).days, 1)

    current_services = calc.get_costs_by_service(start, end)
    previous_services = calc.get_costs_by_service(prev_start, prev_end)
    current_rows = calc.get_costs_by_resource(start, end, include_skus=True)
    previous_rows = calc.get_costs_by_resource(prev_start, prev_end, include_skus=True)
    prev_map = {r.get("resource_id"): float(r.get("total_cost") or 0.0) for r in previous_rows}

    resources = {r.ocid: r for r in db.query(Resource).all()}
    compartments = {c.id: c.name for c in db.query(Compartment).all()}
    latest_scan = db.query(ScanRun).order_by(ScanRun.id.desc()).first()
    settings_snapshot = db.query(Setting).filter(Setting.id == 1).one_or_none()
    app_cfg = get_app_settings()

    current_total = float(sum(current_services.values()))
    previous_total = float(sum(previous_services.values()))
    delta_total = current_total - previous_total

    resource_view: list[dict[str, Any]] = []
    low_confidence_count = 0
    unallocated_count = 0
    rules = load_enabled_rules(db)
    for row in current_rows:
        rid = row.get("resource_id")
        resource = resources.get(rid)
        details = resource.details if resource and resource.details else {}
        compartment_name = compartments.get(
            resource.compartment_id if resource and resource.compartment_id else (row.get("compartment_id") or "unknown"),
            row.get("compartment_name") or "Unknown",
        )
        allocation = evaluate_allocation(
            resource,
            rules,
            compartment_name=compartment_name,
            sku_text=" ".join((s.get("sku_name") or "") for s in (row.get("skus") or [])),
        )
        team = allocation.team
        app = allocation.app
        env = allocation.env
        confidence = allocation.allocation_confidence
        reason = allocation.allocation_reason
        if confidence == "low":
            low_confidence_count += 1
        if team == "Unallocated" or app == "Unallocated" or env == "Unallocated":
            unallocated_count += 1
        resource_view.append(
            {
                "resource_id": rid,
                "name": _safe_name(resource, rid),
                "type": resource.type if resource and resource.type else "unknown",
                "compartment_id": resource.compartment_id if resource and resource.compartment_id else (row.get("compartment_id") or "unknown"),
                "compartment_name": compartment_name,
                "team": team,
                "application": app,
                "environment": env,
                "current_cost": float(row.get("total_cost") or 0.0),
                "previous_cost": float(prev_map.get(rid, 0.0)),
                "skus": row.get("skus") or [],
                "details": details,
                "match_confidence": confidence,
                "match_reason": reason,
            }
        )

    service_items = []
    for name in sorted(set(current_services.keys()) | set(previous_services.keys()), key=lambda n: float(current_services.get(n, 0.0)), reverse=True):
        cur = float(current_services.get(name, 0.0))
        prev = float(previous_services.get(name, 0.0))
        delta = cur - prev
        service_items.append(
            {
                "name": name,
                "current": round(cur, 4),
                "previous": round(prev, 4),
                "delta_abs": round(delta, 4),
                "delta_pct": _safe_pct_rounded(delta, prev),
                "share_pct": _safe_pct_rounded(cur, current_total),
            }
        )

    top_n = int(options.get("top_n", 8))
    min_share_pct = float(options.get("min_share_pct", 0.5))
    tables: dict[str, dict[str, Any]] = {}
    warnings: list[str] = []
    if current_total == 0.0:
        warnings.append("No cost data for selected range")
    if not compartments:
        warnings.append("No compartment tree loaded")

    if report_type in {"executive_summary_monthly", "cost_by_service"}:
        visible = service_items[:top_n]
        if report_type == "cost_by_service":
            remaining = service_items[top_n:]
            if remaining:
                other_cur = sum(i["current"] for i in remaining)
                other_prev = sum(i["previous"] for i in remaining)
                other_delta = other_cur - other_prev
                visible.append(
                    {
                        "name": "Other",
                        "current": round(other_cur, 4),
                        "previous": round(other_prev, 4),
                        "delta_abs": round(other_delta, 4),
                        "delta_pct": _safe_pct_rounded(other_delta, other_prev),
                        "share_pct": _safe_pct_rounded(other_cur, current_total),
                    }
                )
            visible = [i for i in visible if i["name"] == "Other" or i["share_pct"] >= min_share_pct]
        tables["Cost by Service"] = {
            "columns": ["service", "current_cost", "previous_cost", "delta_abs", "delta_pct", "share_pct"],
            "rows": [{"service": i["name"], "current_cost": i["current"], "previous_cost": i["previous"], "delta_abs": i["delta_abs"], "delta_pct": i["delta_pct"], "share_pct": i["share_pct"]} for i in visible],
            "add_totals": True,
        }
        if report_type == "executive_summary_monthly":
            movers = sorted(
                [{"entity_name": r["name"], "entity_type": r["type"], "compartment_name": r["compartment_name"], "current_cost": round(r["current_cost"], 4), "previous_cost": round(r["previous_cost"], 4), "delta_abs": round(r["current_cost"] - r["previous_cost"], 4), "delta_pct": _safe_pct_rounded(r["current_cost"] - r["previous_cost"], r["previous_cost"])} for r in resource_view],
                key=lambda x: abs(x["delta_abs"]),
                reverse=True,
            )[:10]
            tables["Top Movers"] = {"columns": ["entity_name", "entity_type", "compartment_name", "current_cost", "previous_cost", "delta_abs", "delta_pct"], "rows": movers, "add_totals": False}
    elif report_type == "cost_by_compartment":
        grouped: dict[str, dict[str, float]] = {}
        for r in resource_view:
            cid = r["compartment_id"]
            slot = grouped.setdefault(cid, {"current": 0.0, "previous": 0.0})
            slot["current"] += r["current_cost"]
            slot["previous"] += r["previous_cost"]
        rows = []
        for cid, vals in sorted(grouped.items(), key=lambda kv: kv[1]["current"], reverse=True):
            delta = vals["current"] - vals["previous"]
            rows.append({"compartment_name": compartments.get(cid, cid), "compartment_id": cid, "current_cost": round(vals["current"], 4), "previous_cost": round(vals["previous"], 4), "delta_abs": round(delta, 4), "delta_pct": _safe_pct_rounded(delta, vals["previous"]), "share_pct": _safe_pct_rounded(vals["current"], current_total)})
        tables["Cost by Compartment"] = {"columns": ["compartment_name", "compartment_id", "current_cost", "previous_cost", "delta_abs", "delta_pct", "share_pct"], "rows": rows, "add_totals": True}
    elif report_type == "top_resources_by_cost":
        rows = sorted(
            [{"resource_name": r["name"], "resource_type": r["type"], "compartment_name": r["compartment_name"], "current_cost": round(r["current_cost"], 4), "previous_cost": round(r["previous_cost"], 4), "delta_abs": round(r["current_cost"] - r["previous_cost"], 4), "delta_pct": _safe_pct_rounded(r["current_cost"] - r["previous_cost"], r["previous_cost"]), "top_sku": (r["skus"][0].get("sku_name") if r["skus"] else None)} for r in resource_view],
            key=lambda x: x["current_cost"],
            reverse=True,
        )[:max(top_n, 10)]
        tables["Top Resources by Cost"] = {"columns": ["resource_name", "resource_type", "compartment_name", "current_cost", "previous_cost", "delta_abs", "delta_pct", "top_sku"], "rows": rows, "add_totals": True}
    elif report_type == "mapping_health_unallocated":
        grouped: dict[tuple[str, str, str, str, str], float] = {}
        for r in resource_view:
            key = (r["team"], r["application"], r["environment"], r["match_confidence"], r["match_reason"])
            grouped[key] = grouped.get(key, 0.0) + r["current_cost"]
        rows = [{"team": team, "application": app, "environment": env, "current_cost": round(cost, 4), "share_pct": _safe_pct_rounded(cost, current_total), "match_confidence": conf, "match_reason": reason} for (team, app, env, conf, reason), cost in sorted(grouped.items(), key=lambda kv: kv[1], reverse=True)]
        tables["Mapping Health"] = {"columns": ["team", "application", "environment", "current_cost", "share_pct", "match_confidence", "match_reason"], "rows": rows, "add_totals": True}
    elif report_type == "showback_team_app_env":
        by_team: dict[str, dict[str, float]] = {}
        by_app: dict[str, dict[str, float]] = {}
        by_env: dict[str, dict[str, float]] = {}
        unowned_by_comp: dict[str, float] = {}
        unowned_by_service: dict[str, float] = {}
        for r in resource_view:
            delta = r["current_cost"] - r["previous_cost"]
            for grouped, key in [
                (by_team, r["team"]),
                (by_app, r["application"]),
                (by_env, r["environment"]),
            ]:
                slot = grouped.setdefault(key, {"current": 0.0, "previous": 0.0})
                slot["current"] += r["current_cost"]
                slot["previous"] += r["previous_cost"]
            if r["team"] == "Unallocated" or r["application"] == "Unallocated" or r["environment"] == "Unallocated":
                unowned_by_comp[r["compartment_name"]] = unowned_by_comp.get(r["compartment_name"], 0.0) + r["current_cost"]
                sku_text = " ".join((s.get("sku_name") or "") for s in r["skus"]).lower()
                service = "Compute"
                if "storage" in sku_text or "backup" in sku_text or "volume" in sku_text or "snapshot" in sku_text:
                    service = "Storage"
                elif "sql" in sku_text or "database" in sku_text:
                    service = "Database"
                elif "network" in sku_text:
                    service = "Network"
                unowned_by_service[service] = unowned_by_service.get(service, 0.0) + r["current_cost"]
        def _rows(grouped: dict[str, dict[str, float]], col_name: str) -> list[dict]:
            out = []
            for key, vals in sorted(grouped.items(), key=lambda kv: kv[1]["current"], reverse=True):
                delta = vals["current"] - vals["previous"]
                out.append(
                    {
                        col_name: key,
                        "current_cost": round(vals["current"], 4),
                        "previous_cost": round(vals["previous"], 4),
                        "delta_abs": round(delta, 4),
                        "delta_pct": _safe_pct_rounded(delta, vals["previous"]),
                        "share_pct": _safe_pct_rounded(vals["current"], current_total),
                    }
                )
            return out
        tables["Cost by Team"] = {"columns": ["team", "current_cost", "previous_cost", "delta_abs", "delta_pct", "share_pct"], "rows": _rows(by_team, "team"), "add_totals": True}
        tables["Cost by App"] = {"columns": ["application", "current_cost", "previous_cost", "delta_abs", "delta_pct", "share_pct"], "rows": _rows(by_app, "application"), "add_totals": True}
        tables["Cost by Env"] = {"columns": ["environment", "current_cost", "previous_cost", "delta_abs", "delta_pct", "share_pct"], "rows": _rows(by_env, "environment"), "add_totals": True}
        tables["Unowned Breakdown"] = {
            "columns": ["dimension", "name", "current_cost"],
            "rows": (
                [{"dimension": "compartment", "name": k, "current_cost": round(v, 4)} for k, v in sorted(unowned_by_comp.items(), key=lambda kv: kv[1], reverse=True)[:10]]
                + [{"dimension": "service", "name": k, "current_cost": round(v, 4)} for k, v in sorted(unowned_by_service.items(), key=lambda kv: kv[1], reverse=True)[:10]]
            ),
            "add_totals": True,
        }
    elif report_type == "inventory_summary_by_compartment":
        comp_rollup: dict[str, dict[str, float]] = {}
        for r in resource_view:
            cid = r["compartment_id"]
            row = comp_rollup.setdefault(cid, {"compute_count": 0.0, "compute_cost_monthly": 0.0, "db_count": 0.0, "db_cost_monthly": 0.0, "storage_volume_count": 0.0, "storage_cost_monthly": 0.0, "backup_count": 0.0, "backup_cost_monthly": 0.0, "license_windows_cost_monthly": 0.0, "license_sql_cost_monthly": 0.0, "unattached_volume_count": 0.0, "unattached_volume_cost_monthly": 0.0, "total_cost_monthly": 0.0})
            rcost = r["current_cost"]
            row["total_cost_monthly"] += rcost
            rtype = (r["type"] or "").lower()
            if "compute" in rtype or "instance" in rtype:
                row["compute_count"] += 1
                row["compute_cost_monthly"] += rcost
            if "db" in rtype or "database" in rtype or "autonomous" in rtype:
                row["db_count"] += 1
                row["db_cost_monthly"] += rcost
            if "volume" in rtype or "object" in rtype or "storage" in rtype:
                row["storage_volume_count"] += 1
                row["storage_cost_monthly"] += rcost
            if str(r["details"].get("attachment_state") or "").upper() == "UNATTACHED":
                row["unattached_volume_count"] += 1
                row["unattached_volume_cost_monthly"] += rcost
            for sku in r["skus"]:
                sku_name = sku.get("sku_name") or ""
                sku_cost = float(sku.get("cost") or 0.0)
                bucket = _license_bucket(sku_name, r["details"].get("image_name") or "", rtype)
                if "backup" in sku_name.lower() or "snapshot" in sku_name.lower():
                    row["backup_count"] += 1
                    row["backup_cost_monthly"] += sku_cost
                if bucket == "windows":
                    row["license_windows_cost_monthly"] += sku_cost
                elif bucket == "sql":
                    row["license_sql_cost_monthly"] += sku_cost
        rows = []
        for cid, vals in sorted(comp_rollup.items(), key=lambda kv: kv[1]["total_cost_monthly"], reverse=True):
            rows.append({"compartment_name": compartments.get(cid, cid), "compartment_id": cid, "compute_count": int(vals["compute_count"]), "compute_cost_monthly": round(vals["compute_cost_monthly"], 4), "compute_cost_daily": _daily(vals["compute_cost_monthly"], days), "db_count": int(vals["db_count"]), "db_cost_monthly": round(vals["db_cost_monthly"], 4), "db_cost_daily": _daily(vals["db_cost_monthly"], days), "storage_volume_count": int(vals["storage_volume_count"]), "storage_cost_monthly": round(vals["storage_cost_monthly"], 4), "storage_cost_daily": _daily(vals["storage_cost_monthly"], days), "backup_count": int(vals["backup_count"]), "backup_cost_monthly": round(vals["backup_cost_monthly"], 4), "backup_cost_daily": _daily(vals["backup_cost_monthly"], days), "license_windows_cost_monthly": round(vals["license_windows_cost_monthly"], 4), "license_windows_cost_daily": _daily(vals["license_windows_cost_monthly"], days), "license_sql_cost_monthly": round(vals["license_sql_cost_monthly"], 4), "license_sql_cost_daily": _daily(vals["license_sql_cost_monthly"], days), "unattached_volume_count": int(vals["unattached_volume_count"]), "unattached_volume_cost_monthly": round(vals["unattached_volume_cost_monthly"], 4), "unattached_volume_cost_daily": _daily(vals["unattached_volume_cost_monthly"], days), "total_cost_monthly": round(vals["total_cost_monthly"], 4)})
        tables["Inventory Summary"] = {"columns": ["compartment_name", "compartment_id", "compute_count", "compute_cost_monthly", "compute_cost_daily", "db_count", "db_cost_monthly", "db_cost_daily", "storage_volume_count", "storage_cost_monthly", "storage_cost_daily", "backup_count", "backup_cost_monthly", "backup_cost_daily", "license_windows_cost_monthly", "license_windows_cost_daily", "license_sql_cost_monthly", "license_sql_cost_daily", "unattached_volume_count", "unattached_volume_cost_monthly", "unattached_volume_cost_daily", "total_cost_monthly"], "rows": rows, "add_totals": True}
        tables["Top 10 Compartments"] = {"columns": ["compartment_name", "compartment_id", "total_cost_monthly"], "rows": rows[:10], "add_totals": True}
    elif report_type == "storage_backup_governance":
        rows = []
        for r in resource_view:
            backup_cost = sum(float(sku.get("cost") or 0.0) for sku in r["skus"] if "backup" in (sku.get("sku_name") or "").lower() or "snapshot" in (sku.get("sku_name") or "").lower())
            rows.append({"compartment_name": r["compartment_name"], "resource_name": r["name"], "resource_type": r["type"], "is_unattached": str(r["details"].get("attachment_state") or "").upper() == "UNATTACHED", "monthly_cost": round(r["current_cost"], 4), "daily_cost": _daily(r["current_cost"], days), "backup_monthly_cost": round(backup_cost, 4), "backup_daily_cost": _daily(backup_cost, days)})
        tables["Storage & Backup Governance"] = {"columns": ["compartment_name", "resource_name", "resource_type", "is_unattached", "monthly_cost", "daily_cost", "backup_monthly_cost", "backup_daily_cost"], "rows": rows, "add_totals": True}
    elif report_type == "license_spend":
        rows = []
        for r in resource_view:
            for sku in r["skus"]:
                sku_name = sku.get("sku_name") or ""
                cost = float(sku.get("cost") or 0.0)
                bucket = _license_bucket(sku_name, r["details"].get("image_name") or "", r["type"])
                if bucket == "other":
                    continue
                rows.append({"license_category": bucket, "resource_name": r["name"], "resource_type": r["type"], "compartment_name": r["compartment_name"], "sku_name": sku_name, "monthly_cost": round(cost, 4), "daily_cost": _daily(cost, days)})
        tables["License Spend"] = {"columns": ["license_category", "resource_name", "resource_type", "compartment_name", "sku_name", "monthly_cost", "daily_cost"], "rows": rows, "add_totals": True}
    elif report_type == "movers_and_anomalies":
        rows = sorted(
            [{"entity_name": r["name"], "entity_type": r["type"], "compartment_name": r["compartment_name"], "current_cost": round(r["current_cost"], 4), "previous_cost": round(r["previous_cost"], 4), "delta_abs": round(r["current_cost"] - r["previous_cost"], 4), "delta_pct": _safe_pct_rounded(r["current_cost"] - r["previous_cost"], r["previous_cost"])} for r in resource_view],
            key=lambda x: abs(x["delta_abs"]),
            reverse=True,
        )[:20]
        tables["Movers & Anomalies"] = {"columns": ["entity_name", "entity_type", "compartment_name", "current_cost", "previous_cost", "delta_abs", "delta_pct"], "rows": rows, "add_totals": False}
    elif report_type == "optimization_recommendations":
        rec_payload = generate_recommendations(db=db, start=start, end_exclusive=end)
        rec_items = rec_payload["items"]
        by_category: dict[str, dict[str, float]] = {}
        high_rows = []
        medlow_rows = []
        unattached_rows = []
        license_rows = []
        for rec in rec_items:
            slot = by_category.setdefault(rec.category, {"count": 0.0, "savings": 0.0})
            slot["count"] += 1.0
            slot["savings"] += rec.estimated_savings
            row = {
                "recommendation_id": rec.recommendation_id,
                "type": rec.type,
                "category": rec.category,
                "resource_name": rec.resource_name,
                "resource_ref": rec.resource_ref,
                "compartment_name": rec.compartment_name,
                "team": rec.team,
                "app": rec.app,
                "env": rec.env,
                "current_cost": round(rec.current_cost, 4),
                "estimated_savings": round(rec.estimated_savings, 4),
                "confidence": rec.confidence,
                "reason": rec.reason,
                "recommendation": rec.recommendation,
            }
            if rec.confidence == "high":
                high_rows.append(row)
            else:
                medlow_rows.append(row)
            if rec.type == "unattached_resource":
                unattached_rows.append(row)
            if rec.category == "license":
                license_rows.append(row)

        summary_rows = [
            {
                "category": category,
                "count": int(vals["count"]),
                "savings_monthly": round(vals["savings"], 4),
            }
            for category, vals in sorted(by_category.items(), key=lambda kv: kv[0])
        ]
        tables["Optimization Summary"] = {
            "columns": ["category", "count", "savings_monthly"],
            "rows": summary_rows,
            "add_totals": True,
        }
        rec_cols = [
            "recommendation_id",
            "type",
            "category",
            "resource_name",
            "resource_ref",
            "compartment_name",
            "team",
            "app",
            "env",
            "current_cost",
            "estimated_savings",
            "confidence",
            "reason",
            "recommendation",
        ]
        tables["High Confidence Recommendations"] = {"columns": rec_cols, "rows": high_rows, "add_totals": False}
        tables["Medium Low Recommendations"] = {"columns": rec_cols, "rows": medlow_rows, "add_totals": False}
        tables["Unattached Resources"] = {"columns": rec_cols, "rows": unattached_rows, "add_totals": False}
        tables["License Signals"] = {"columns": rec_cols, "rows": license_rows, "add_totals": False}
        if not rec_items:
            warnings.append("No optimization opportunities detected for selected range")
    elif report_type == "budget_health":
        statuses = evaluate_budget_statuses(db, persist_alerts=True)
        summary_rows = [
            {
                "budget_id": s.budget_id,
                "budget_name": s.budget_name,
                "scope_type": s.scope_type,
                "scope_value": s.scope_value,
                "current_spend": round(s.current_spend, 4),
                "budget_limit": round(s.budget_limit, 4),
                "utilization_pct": round(s.utilization_pct, 4),
                "forecast_end_of_month": round(s.forecast_end_of_month, 4),
                "breach_level": s.breach_level,
                "days_remaining": s.days_remaining,
                "explanation": s.explanation,
                "forecast_narrative": getattr(s, "narrative", ""),
            }
            for s in statuses
        ]
        breached = [r for r in summary_rows if r["breach_level"] == "critical"]
        risks = [r for r in summary_rows if r["forecast_end_of_month"] >= r["budget_limit"] and r["breach_level"] != "critical"]
        alert_rows = []
        alert_events = db.query(BudgetAlertEvent).order_by(BudgetAlertEvent.triggered_at.desc()).limit(200).all()
        for event in alert_events:
            payload = event.payload or {}
            alert_rows.append(
                {
                    "triggered_at": event.triggered_at.isoformat().replace("+00:00", "Z") if event.triggered_at else None,
                    "budget_id": event.budget_id,
                    "alert_kind": event.alert_kind,
                    "threshold": event.threshold,
                    "reason": payload.get("reason") or "",
                    "current_spend": float(payload.get("current_spend") or 0.0),
                    "projected_spend": float(payload.get("projected_spend") or 0.0),
                }
            )
        exec_summary = [
            {
                "metric": "total_budgets",
                "value": len(summary_rows),
            },
            {
                "metric": "breached_budgets",
                "value": len(breached),
            },
            {
                "metric": "forecast_risk_budgets",
                "value": len(risks),
            },
            {
                "metric": "alerts_in_timeline",
                "value": len(alert_rows),
            },
        ]
        history_rows = [
            {
                "budget_id": s.budget_id,
                "budget_name": s.budget_name,
                "current_spend": round(s.current_spend, 4),
                "forecast_end_of_month": round(s.forecast_end_of_month, 4),
                "utilization_pct": round(s.utilization_pct, 4),
                "generated_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
            }
            for s in statuses
        ]
        cols = [
            "budget_id",
            "budget_name",
            "scope_type",
            "scope_value",
            "current_spend",
            "budget_limit",
            "utilization_pct",
            "forecast_end_of_month",
            "breach_level",
            "days_remaining",
            "explanation",
            "forecast_narrative",
        ]
        tables["Executive Summary"] = {"columns": ["metric", "value"], "rows": exec_summary, "add_totals": False}
        tables["Budget Summary"] = {"columns": cols, "rows": summary_rows, "add_totals": False}
        tables["Breached Budgets"] = {"columns": cols, "rows": breached, "add_totals": False}
        tables["Forecast Risks"] = {"columns": cols, "rows": risks, "add_totals": False}
        tables["Alert Timeline"] = {
            "columns": ["triggered_at", "budget_id", "alert_kind", "threshold", "reason", "current_spend", "projected_spend"],
            "rows": alert_rows,
            "add_totals": False,
        }
        tables["Budget History Snapshot"] = {
            "columns": ["budget_id", "budget_name", "current_spend", "forecast_end_of_month", "utilization_pct", "generated_at"],
            "rows": history_rows,
            "add_totals": False,
        }
        if not statuses:
            warnings.append("No budgets configured")
    elif report_type == "actions_audit":
        actions = db.query(ActionRequest).order_by(ActionRequest.created_at.desc()).all()
        events = db.query(ActionEvent).order_by(ActionEvent.timestamp.desc(), ActionEvent.id.desc()).all()
        summary_rows = [
            {
                "action_id": a.action_id,
                "source": a.source,
                "category": a.category,
                "target_type": a.target_type,
                "status": a.status,
                "risk_level": a.risk_level,
                "confidence": a.confidence,
                "estimated_savings_monthly": round(float(a.estimated_savings_monthly or 0.0), 4),
                "requested_by": a.requested_by,
                "approved_by": a.approved_by,
                "created_at": a.created_at.isoformat().replace("+00:00", "Z") if a.created_at else None,
                "updated_at": a.updated_at.isoformat().replace("+00:00", "Z") if a.updated_at else None,
            }
            for a in actions
        ]
        timeline_rows = [
            {
                "action_id": e.action_id,
                "event_type": e.event_type,
                "message": e.message,
                "timestamp": e.timestamp.isoformat().replace("+00:00", "Z") if e.timestamp else None,
            }
            for e in events
        ]
        savings_rows = [
            {
                "action_id": a.action_id,
                "estimated_savings_monthly": round(float(a.estimated_savings_monthly or 0.0), 4),
                "actual_savings_monthly": 0.0,
                "status": a.status,
            }
            for a in actions
        ]
        tables["Action Summary"] = {
            "columns": [
                "action_id",
                "source",
                "category",
                "target_type",
                "status",
                "risk_level",
                "confidence",
                "estimated_savings_monthly",
                "requested_by",
                "approved_by",
                "created_at",
                "updated_at",
            ],
            "rows": summary_rows,
            "add_totals": False,
        }
        tables["Action Timeline"] = {
            "columns": ["action_id", "event_type", "message", "timestamp"],
            "rows": timeline_rows,
            "add_totals": False,
        }
        tables["Savings Realization"] = {
            "columns": ["action_id", "estimated_savings_monthly", "actual_savings_monthly", "status"],
            "rows": savings_rows,
            "add_totals": True,
        }
    elif report_type == "ops_audit":
        scans = db.query(ScanRun).order_by(ScanRun.started_at.desc()).limit(200).all()
        alerts = db.query(BudgetAlertEvent).order_by(BudgetAlertEvent.triggered_at.desc()).limit(200).all()
        actions = db.query(ActionRequest).order_by(ActionRequest.created_at.desc()).limit(500).all()
        scan_rows = [
            {
                "scan_id": s.id,
                "status": s.status,
                "started_at": s.started_at.isoformat().replace("+00:00", "Z") if s.started_at else None,
                "finished_at": s.finished_at.isoformat().replace("+00:00", "Z") if s.finished_at else None,
                "error_message": s.error_message or "",
            }
            for s in scans
        ]
        alert_rows = [
            {
                "alert_id": a.id,
                "budget_id": a.budget_id,
                "alert_kind": a.alert_kind,
                "threshold": a.threshold,
                "triggered_at": a.triggered_at.isoformat().replace("+00:00", "Z") if a.triggered_at else None,
            }
            for a in alerts
        ]
        status_rollup: dict[str, int] = {}
        failure_rollup: dict[str, int] = {}
        for a in actions:
            status_rollup[a.status] = status_rollup.get(a.status, 0) + 1
        for e in db.query(ActionEvent).filter(ActionEvent.event_type == "failed").all():
            reason = (e.message or "unknown")[:120]
            failure_rollup[reason] = failure_rollup.get(reason, 0) + 1
        action_rows = [{"status": k, "count": v} for k, v in sorted(status_rollup.items(), key=lambda kv: kv[0])]
        failure_rows = [{"failure_reason": k, "count": v} for k, v in sorted(failure_rollup.items(), key=lambda kv: kv[1], reverse=True)[:20]]
        tables["Scans Timeline"] = {"columns": ["scan_id", "status", "started_at", "finished_at", "error_message"], "rows": scan_rows, "add_totals": False}
        tables["Alerts Timeline"] = {"columns": ["alert_id", "budget_id", "alert_kind", "threshold", "triggered_at"], "rows": alert_rows, "add_totals": False}
        tables["Actions By Status"] = {"columns": ["status", "count"], "rows": action_rows, "add_totals": True}
        tables["Failures Root Causes"] = {"columns": ["failure_reason", "count"], "rows": failure_rows, "add_totals": True}
    elif report_type == "vm_windows_inventory":
        _VM_TYPES = {"compute", "instance", "windows_server", "windows"}
        vm_rows: list[dict] = []
        for r in sorted(resources.values(), key=lambda x: (x.name or "").lower()):
            rtype_lower = (r.type or "").lower()
            is_vm = any(t in rtype_lower for t in _VM_TYPES) or bool(r.shape)
            if not is_vm:
                continue
            details = r.details or {}
            comp_name = compartments.get(r.compartment_id, r.compartment_id or "")
            alloc = evaluate_allocation(
                r, rules, compartment_name=comp_name, sku_text=""
            )
            # Private IP — stored as list or string
            raw_ips = details.get("private_ips") or details.get("private_ip") or []
            if isinstance(raw_ips, list):
                private_ip = ", ".join(str(ip) for ip in raw_ips if ip)
            else:
                private_ip = str(raw_ips) if raw_ips else ""
            # OCPUs
            ocpus_raw = (
                details.get("ocpus")
                or details.get("ocpu_count")
                or details.get("cpu_count")
                or ""
            )
            ocpus = str(ocpus_raw) if ocpus_raw != "" else ""
            # Memory
            mem_raw = details.get("memory_in_gbs") or details.get("memory_gb") or ""
            memory = f"{mem_raw} GB" if mem_raw else ""
            vm_rows.append({
                "name": r.name or "",
                "type": r.type or "compute",
                "status": (r.status or "").upper(),
                "compartment": comp_name,
                "env": alloc.env or "Unallocated",
                "team": alloc.team or "Unallocated",
                "app": alloc.app or "Unallocated",
                "confidence": alloc.allocation_confidence or "low",
                "shape": r.shape or "",
                "ocpus": ocpus,
                "memory": memory,
                "private_ip": private_ip,
                "image": details.get("image_name") or details.get("image") or "",
            })
        tables["VM & Windows Server Inventory"] = {
            "columns": [
                "name", "type", "status", "compartment",
                "env", "team", "app", "confidence",
                "shape", "ocpus", "memory", "private_ip", "image",
            ],
            "rows": vm_rows,
            "add_totals": False,
        }
    else:
        raise HTTPException(status_code=400, detail="Unknown report_type")

    if unallocated_count > 0:
        warnings.append("Unallocated owner/team/app tags detected")
    if low_confidence_count > 0:
        warnings.append("Low-confidence mapping entries detected")

    summary = {
        "title": _report_catalog()[report_type]["title"],
        "report_type": report_type,
        "start_date": start_date_iso,
        "end_date": end_date_iso,
        "generated_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "generated_by": user_name,
        "compare": compare,
        "total_current_cost": round(current_total, 4),
        "total_previous_cost": round(previous_total, 4),
        "delta_abs": round(delta_total, 4),
        "delta_pct": _safe_pct_rounded(delta_total, previous_total),
        "warnings": "; ".join(warnings) if warnings else "",
    }
    manifest = {
        "report_type": report_type,
        "start_date": start_date_iso,
        "end_date": end_date_iso,
        "compare": compare,
        "selected_filters": {
            "group_by": options.get("group_by"),
            "top_n": top_n,
            "min_share_pct": min_share_pct,
            "include_children": bool(options.get("include_children", True)),
            "compartment_ids": list(options.get("compartment_ids") or []),
        },
        "generated_at": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "generated_by": user_name,
        "actor": user_name,
        "app_version": getattr(app_cfg, "app_version", None),
        "git_commit": _get_git_commit(),
        "scan_id": latest_scan.id if latest_scan else None,
        "scan_run_id": latest_scan.id if latest_scan else None,
        "last_scan_at": latest_scan.finished_at.isoformat().replace("+00:00", "Z") if latest_scan and latest_scan.finished_at else None,
        "oci_auth_mode": getattr(settings_snapshot, "oci_auth_mode", "profile"),
        "oci_config_profile": getattr(settings_snapshot, "oci_config_profile", "DEFAULT"),
    }
    if report_type == "optimization_recommendations":
        manifest["detection_rules_used"] = [
            "unattached_block_or_boot_volume",
            "orphaned_backup_retention_gt_30_days",
            "idle_or_underutilized_compute_heuristic",
            "oversized_storage_threshold_gte_500gb",
            "license_windows_sql_signal",
        ]
        manifest["confidence_criteria"] = {
            "high": "deterministic state (unattached/retention matched)",
            "medium": "heuristic utilization/size evidence",
            "low": "license advisory signal",
        }
    if report_type == "budget_health":
        manifest["evaluation_date"] = datetime.now(UTC).isoformat().replace("+00:00", "Z")
        manifest["thresholds"] = [50, 75, 90, 100]
        manifest["forecast_method"] = "linear_daily_burn_rate_to_end_of_month"
        try:
            setting = db.query(Setting).filter(Setting.id == 1).one_or_none()
        except Exception:
            setting = None
        manifest["notification_channels_enabled"] = {
            "email": bool(getattr(setting, "notifications_email_enabled", False)) if setting else False,
            "webhook": bool(getattr(setting, "notifications_webhook_enabled", False)) if setting else False,
        }
        manifest["narrative_rules_version"] = "v1"
    if report_type == "actions_audit":
        manifest["approval_policy"] = "explicit_admin_approval_required_unless_risk_level_safe"
        manifest["executor_versions"] = {
            "notify_only": "v1",
            "cleanup_unattached_volume": "v1",
            "stop_idle_instance": "v1",
            "tag_fix": "v1",
        }
    if report_type == "ops_audit":
        s = db.query(Setting).filter(Setting.id == 1).one_or_none()
        manifest["build_version"] = getattr(app_cfg, "app_version", "unknown")
        manifest["feature_flags_snapshot"] = feature_flags(s)
    validation = {
        "row_count": {name: len(table["rows"]) for name, table in tables.items()},
        "totals_checksum": {
            "current_total_cost": round(current_total, 6),
            "previous_total_cost": round(previous_total, 6),
        },
        "unmapped_unallocated_pct": round((unallocated_count / max(len(resource_view), 1)) * 100.0, 4),
        "low_confidence_count": low_confidence_count,
        "warnings": warnings,
    }
    if report_type == "optimization_recommendations":
        savings_values = [
            float(row.get("estimated_savings") or 0.0)
            for row in tables.get("High Confidence Recommendations", {}).get("rows", [])
        ] + [
            float(row.get("estimated_savings") or 0.0)
            for row in tables.get("Medium Low Recommendations", {}).get("rows", [])
        ]
        validation["savings_non_negative"] = all(v >= 0 for v in savings_values)
    if report_type == "budget_health":
        validation["savings_non_negative"] = True
    if report_type == "actions_audit":
        allowed_statuses = {
            "draft",
            "pending_approval",
            "approved",
            "rejected",
            "queued",
            "running",
            "succeeded",
            "failed",
            "rolled_back",
        }
        status_values = [str(r.get("status") or "") for r in tables.get("Action Summary", {}).get("rows", [])]
        validation["status_transitions_valid"] = all(s in allowed_statuses for s in status_values)
    return {"summary": summary, "tables": tables, "manifest": manifest, "validation": validation}


def _write_catalog_xlsx(path: Path, payload: dict) -> None:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["key", "value"])
    for k, v in payload.get("summary", {}).items():
        ws_summary.append([k, v])
    header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
    for cell in ws_summary[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    ws_summary.freeze_panes = "A2"

    for sheet_key, table in payload.get("tables", {}).items():
        ws = wb.create_sheet(_sheet_name(sheet_key))
        headers = table.get("columns") or []
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for row in table.get("rows", []):
            ws.append([row.get(h) for h in headers])

        if table.get("add_totals") and table.get("rows"):
            totals = []
            for idx, h in enumerate(headers):
                values = [r.get(h) for r in table["rows"]]
                if all(isinstance(v, (int, float)) for v in values if v is not None):
                    totals.append(round(sum(float(v or 0.0) for v in values), 4))
                else:
                    totals.append("TOTAL" if idx == 0 else None)
            ws.append(totals)
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)

        ws.freeze_panes = "A2"
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 48)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for idx, cell in enumerate(row):
                head = headers[idx].lower() if idx < len(headers) else ""
                if isinstance(cell.value, (int, float)):
                    if "pct" in head or "share" in head:
                        cell.number_format = "0.00%"
                        cell.value = float(cell.value) / 100.0
                    elif any(k in head for k in ["cost", "monthly", "daily", "total", "delta", "checksum"]):
                        cell.number_format = "$#,##0.00"
    wb.save(path)


@router.post("/exports/generate")
async def generate_report(req: ExportGenerateRequest, db: Session = Depends(get_db), user=Depends(_require_admin)):
    catalog = _report_catalog()
    report_type = _normalize_report_type(req.report_type)
    if report_type not in catalog:
        raise HTTPException(status_code=400, detail=f"Unknown report_type '{req.report_type}'")
    compare = str((req.options or {}).get("compare") or "previous")
    if compare != "previous":
        raise HTTPException(status_code=422, detail="compare must be 'previous'")
    start = _parse_iso_date(req.start_date, is_end=False)
    end = _parse_iso_date(req.end_date, is_end=True)
    if end <= start:
        raise HTTPException(status_code=422, detail="end_date must be after start_date")
    start_iso = start.date().isoformat()
    end_iso = (end - timedelta(days=1)).date().isoformat()

    report = _build_report_data_v2(report_type, start, end, start_iso, end_iso, compare, req.options or {}, db, user.get("sub"))
    now = datetime.now(UTC)
    stamp = now.strftime("%Y%m%d-%H%M%S")
    cfg = get_app_settings()
    export_dir = Path(cfg.export_dir)
    export_dir.mkdir(parents=True, exist_ok=True)
    base = f"report-{report_type}-{stamp}"
    xlsx_name = f"{base}.xlsx"
    manifest_name = f"{base}.manifest.json"
    validation_name = f"{base}.validation.json"
    _write_catalog_xlsx(export_dir / xlsx_name, report)
    with (export_dir / manifest_name).open("w", encoding="utf-8") as mf:
        json.dump(report["manifest"], mf, indent=2)
    with (export_dir / validation_name).open("w", encoding="utf-8") as vf:
        json.dump(report["validation"], vf, indent=2)
    return {
        "success": True,
        "data": {
            "files": {
                "xlsx": {
                    "name": xlsx_name,
                    "download_url": f"/api/v1/admin/exports/download/{xlsx_name}",
                },
                "manifest": {
                    "name": manifest_name,
                    "download_url": f"/api/v1/admin/exports/download/{manifest_name}",
                },
                "validation": {
                    "name": validation_name,
                    "download_url": f"/api/v1/admin/exports/download/{validation_name}",
                },
            }
        },
    }


@router.get("/users")
async def list_users(db: Session = Depends(get_db), user=Depends(_require_admin)):
    rows = db.query(UserAccount).order_by(UserAccount.username.asc()).all()
    return {
        "success": True,
        "data": [
            {
                "id": r.id,
                "username": r.username,
                "role": r.role,
                "allowed_teams": list(r.allowed_teams or []),
                "allowed_apps": list(r.allowed_apps or []),
                "allowed_envs": list(r.allowed_envs or []),
                "allowed_compartment_ids": list(r.allowed_compartment_ids or []),
                "is_active": bool(r.is_active),
                "created_at": r.created_at.isoformat() if r.created_at else None,
                "updated_at": r.updated_at.isoformat() if r.updated_at else None,
            }
            for r in rows
        ],
    }


@router.post("/users")
async def create_user(req: UserCreateRequest, db: Session = Depends(get_db), user=Depends(_require_admin)):
    username = (req.username or "").strip()
    if not username:
        raise HTTPException(status_code=400, detail="username is required")
    if db.query(UserAccount).filter(UserAccount.username == username).one_or_none():
        raise HTTPException(status_code=409, detail="username already exists")
    role = (req.role or "viewer").strip().lower()
    if role not in {"admin", "finops", "engineer", "viewer"}:
        raise HTTPException(status_code=400, detail="invalid role")
    row = UserAccount(
        username=username,
        password_hash=hash_password(req.password),
        role=role,
        allowed_teams=[x.strip() for x in (req.allowed_teams or []) if x and x.strip()],
        allowed_apps=[x.strip() for x in (req.allowed_apps or []) if x and x.strip()],
        allowed_envs=[x.strip() for x in (req.allowed_envs or []) if x and x.strip()],
        allowed_compartment_ids=[x.strip() for x in (req.allowed_compartment_ids or []) if x and x.strip()],
        is_active=bool(req.is_active),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    return {"success": True, "data": {"id": row.id, "username": row.username}}


@router.put("/users/{user_id}")
async def update_user(user_id: int, req: UserUpdateRequest, db: Session = Depends(get_db), user=Depends(_require_admin)):
    row = db.query(UserAccount).filter(UserAccount.id == user_id).one_or_none()
    if not row:
        raise HTTPException(status_code=404, detail="user not found")
    if req.password:
        row.password_hash = hash_password(req.password)
    if req.role is not None:
        role = (req.role or "").strip().lower()
        if role not in {"admin", "finops", "engineer", "viewer"}:
            raise HTTPException(status_code=400, detail="invalid role")
        row.role = role
    if req.allowed_teams is not None:
        row.allowed_teams = [x.strip() for x in (req.allowed_teams or []) if x and x.strip()]
    if req.allowed_apps is not None:
        row.allowed_apps = [x.strip() for x in (req.allowed_apps or []) if x and x.strip()]
    if req.allowed_envs is not None:
        row.allowed_envs = [x.strip() for x in (req.allowed_envs or []) if x and x.strip()]
    if req.allowed_compartment_ids is not None:
        row.allowed_compartment_ids = [x.strip() for x in (req.allowed_compartment_ids or []) if x and x.strip()]
    if req.is_active is not None:
        if row.role == "admin" and req.is_active is False:
            active_admins = db.query(UserAccount).filter(UserAccount.role == "admin", UserAccount.is_active == True).count()
            if active_admins <= 1:
                raise HTTPException(status_code=400, detail="cannot deactivate last active admin")
        row.is_active = bool(req.is_active)
    db.commit()
    return {"success": True}
